import os
import re
from datetime import datetime, date
from shutil import copyfile
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import streamlit as st

# ---------------- CONFIG ----------------
HEADER_ROW = 2        # green header row in template (dates + สาย + summary headers)
ID_COL = 2            # column B in template (employee ID, 1-based index)
FIRST_DATE_COL = 4    # column D in template (first date column, 1-based index)


# ----------------------------------------
# Helpers (same logic as your working debug.py)
# ----------------------------------------

def parse_date_cell(x):
    """
    Convert various Excel / string values into a Python date.
    NO conversion between BE/CE: if the sheet says 2568, we keep 2568.
    """
    if isinstance(x, (datetime, date)):
        return x.date()

    if x is None:
        return None

    s = str(x).strip()
    if not s:
        return None

    # explicit dd/mm/yyyy
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        d, m, y = map(int, s.split("/"))
        return date(y, m, d)

    # fallback: let pandas try
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(dt):
            return None
        dt = dt.to_pydatetime()
        return dt.date()
    except Exception:
        return None


def load_day_file(path: str) -> pd.DataFrame:
    """
    Read the multi-day file (e.g. '26.11-25.12.2568.xlsx').

    Cols (0-based):
      0 = emp_id (first row of block, then blank) -> forward-filled
      1 = name   (same)
      2 = date string (26/11/2568, etc)
      4 = time_in
      5 = time_out
      6 = late minutes
      9..15 = reason counts (ไม่ตอกเข้า, ไม่ตอกออก, ขาดงาน, ลาป่วย, ลากิจ, พักร้อน, บวชคลอด)
      16 = comment (Q)
    """
    print(f"Loading day file: {path}")
    df = pd.read_excel(path, header=None)

    # forward fill emp_id and name downwards
    df[0] = df[0].ffill()
    df[1] = df[1].ffill()

    # parse dates
    df["date_raw"] = df[2]
    df["date"] = df["date_raw"].apply(parse_date_cell)

    print("  sample raw date values (first 10):", list(df["date_raw"].head(10)))
    unique_dates = sorted({d for d in df["date"] if isinstance(d, date)})
    print("  parsed unique dates:", unique_dates)

    # keep only rows that have a valid date
    mask = df["date"].notna()
    df_emp = df.loc[mask].copy()

    print("  rows kept (with valid date):", len(df_emp))
    print("  unique employees in file:", df_emp[0].nunique())

    # rename useful columns
    rename_map = {
        0: "emp_id",
        1: "name",
        2: "date_raw",
        4: "time_in",
        5: "time_out",
    }

    if 6 in df_emp.columns:   # late
        rename_map[6] = "late"

    # J..P reasons
    if 9 in df_emp.columns:   rename_map[9]  = "no_in"
    if 10 in df_emp.columns:  rename_map[10] = "no_out"
    if 11 in df_emp.columns:  rename_map[11] = "absent"
    if 12 in df_emp.columns:  rename_map[12] = "sick"
    if 13 in df_emp.columns:  rename_map[13] = "personal"
    if 14 in df_emp.columns:  rename_map[14] = "vacation"
    if 15 in df_emp.columns:  rename_map[15] = "ordination"

    # Q column (index 16) = comment
    if 16 in df_emp.columns:
        rename_map[16] = "comment"

    df_emp = df_emp.rename(columns=rename_map)

    df_emp["emp_id"] = df_emp["emp_id"].astype(str).str.strip()
    df_emp["name"] = df_emp["name"].astype(str).str.strip()

    # for debug: show some comment rows
    if "comment" in df_emp.columns:
        sample_comments = df_emp[["emp_id", "date", "time_in", "comment"]].dropna(subset=["comment"]).head()
        print("  sample rows with comment (up to 5):")
        print(sample_comments)

    return df_emp


def fill_template_from_days(template_path: str, day_files: list[str], output_path: str):
    """
    Copy template.xlsx to output_path and fill:
      * per-day time_in in date columns
      * yellow highlight + text (reasons or comments) when no time_in
      * accumulate late minutes into 'สาย (นาที)'
      * final summary columns: ขาดงาน / ลาป่วย / ลากิจ / พักร้อน / บวชคลอด
    """
    # delete old output
    if os.path.exists(output_path):
        os.remove(output_path)

    copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # ---------- 1) Build date_col_map and find 'สาย (นาที)' + summary columns ----------
    date_col_map = {}     # {date: col_index}
    late_col_idx = None

    summary_cols = {
        "absent": None,
        "sick": None,
        "personal": None,
        "vacation": None,
        "ordination": None,
    }

    max_col = ws.max_column

    for c in range(FIRST_DATE_COL, max_col + 1):
        header_val = ws.cell(row=HEADER_ROW, column=c).value
        if header_val is None:
            continue
        d = parse_date_cell(header_val)
        if isinstance(d, date):
            date_col_map[d] = c

        text = str(header_val)
        if late_col_idx is None and ("สาย" in text or "late" in text.lower()):
            late_col_idx = c

    # also search for summary headers on the right
    for c in range(1, max_col + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v is None:
            continue
        text = str(v)
        if "ขาดงาน" in text:
            summary_cols["absent"] = c
        elif "ลาป่วย" in text:
            summary_cols["sick"] = c
        elif "ลากิจ" in text:
            summary_cols["personal"] = c
        elif "พักร้อน" in text:
            summary_cols["vacation"] = c
        elif "บวช" in text:
            summary_cols["ordination"] = c

    print("Date headers found in template:")
    for d, col in date_col_map.items():
        print(f"  {d} -> col {col}")
    print("Late column index in template:", late_col_idx)
    print("Summary columns:", summary_cols)

    if not date_col_map:
        raise RuntimeError("ไม่พบหัวคอลัมน์วันที่ใน template")

    # ---------- 2) Index employees by ID ----------
    index_by_id = {}
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        emp_id_val = ws.cell(row=r, column=ID_COL).value
        if emp_id_val is None:
            continue
        emp_id = str(emp_id_val).strip()
        if emp_id:
            index_by_id[emp_id] = r

    print(f"Number of employees in template (unique IDs): {len(index_by_id)}")

    # ---------- 3) Fill per-day values ----------
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    total_matches = 0

    for path in day_files:
        df = load_day_file(path)

        for _, row in df.iterrows():
            emp_id = str(row["emp_id"]).strip()
            d = row["date"]
            if not isinstance(d, date):
                continue

            col_idx = date_col_map.get(d)
            if col_idx is None:
                # date not in template
                continue

            row_idx = index_by_id.get(emp_id)
            if row_idx is None:
                # employee not in template
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            time_in = row.get("time_in", None)

            # ---------- choose text for that date cell ----------
            if pd.notna(time_in):
                # normal case: has time-in
                cell.value = str(time_in)
                total_matches += 1
            else:
                # no time-in: highlight yellow and add text
                cell.fill = yellow_fill

                # 1) comment overrides everything
                comment_str = None
                if "comment" in row.index and pd.notna(row["comment"]):
                    comment_str = str(row["comment"]).strip()
                    if not comment_str:
                        comment_str = None

                label_map = {
                    "no_in":      "ไม่ตอกเข้า",
                    "no_out":     "ไม่ตอกออก",
                    "absent":     "ขาดงาน",
                    "sick":       "ลาป่วย",
                    "personal":   "ลากิจ",
                    "vacation":   "พักร้อน",
                    "ordination": "บวชคลอด",
                }

                pieces = []
                if comment_str is None:
                    # build from numeric reason columns
                    for key, label in label_map.items():
                        if key not in row.index:
                            continue
                        val = row[key]
                        if pd.isna(val):
                            continue
                        try:
                            num = float(val)
                        except (TypeError, ValueError):
                            continue
                        if num == 0:
                            continue

                        if abs(num - round(num)) < 1e-6:
                            num_str = str(int(round(num)))
                        else:
                            num_str = str(num)

                        if num_str == "1":
                            pieces.append(label)
                        else:
                            pieces.append(f"{label}({num_str})")

                # choose final cell text
                if comment_str:
                    cell.value = comment_str
                elif pieces:
                    cell.value = " ".join(pieces)
                else:
                    # no comment & no specific status -> treat as ขาดงาน
                    cell.value = "ขาดงาน"

            # ---------- accumulate late minutes ----------
            if late_col_idx is not None and "late" in row.index:
                late_val = row["late"]
                if pd.notna(late_val):
                    try:
                        add_val = float(late_val)
                    except (TypeError, ValueError):
                        add_val = 0.0
                    if add_val != 0.0:
                        existing = ws.cell(row=row_idx, column=late_col_idx).value
                        try:
                            base = float(existing) if existing is not None else 0.0
                        except (TypeError, ValueError):
                            base = 0.0
                        ws.cell(row=row_idx, column=late_col_idx).value = base + add_val

    print("Total cells filled with time_in:", total_matches)

    # ---------- 4) Summary counts per person ----------
    if any(summary_cols.values()):
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            counts = {k: 0 for k in summary_cols.keys()}

            for col_idx in date_col_map.values():
                v = ws.cell(row=r, column=col_idx).value
                if v is None:
                    continue
                s = str(v)

                if "ขาดงาน" in s:
                    counts["absent"] += 1
                if "ป่วย" in s:
                    counts["sick"] += 1
                if "กิจ" in s:
                    counts["personal"] += 1
                if "พักร้อน" in s:
                    counts["vacation"] += 1
                if "บวช" in s:
                    counts["ordination"] += 1

            for key, col in summary_cols.items():
                if col is not None:
                    ws.cell(row=r, column=col).value = counts[key]

    wb.save(output_path)
    print("\nSaved filled template to:", output_path)


# ----------------- STREAMLIT UI -----------------

st.set_page_config(page_title="สรุปการมาทำงาน", layout="wide")

st.title("สรุปการมาทำงาน (Time Attendance Helper)")

# --- TEMPLATE DOWNLOAD SECTION ---
st.subheader("📄 Template ตัวอย่าง (ต้องใช้รูปแบบนี้เท่านั้น)")
template_public_path = os.path.join(os.path.dirname(__file__), "template-public.xlsx")

if os.path.exists(template_public_path):
    with open(template_public_path, "rb") as f:
        st.download_button(
            label="📥 ดาวน์โหลด template ตัวอย่าง",
            data=f,
            file_name="template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.warning("ไม่พบไฟล์ template-public.xlsx ในโฟลเดอร์เดียวกับ app.py")

st.markdown("---")

st.markdown(
    """
    ### วิธีใช้งาน

    1. อัปโหลดไฟล์ **template** (สรุปการทำงานประจำงวด) – ต้องเป็น `.xlsx`  
    2. อัปโหลดไฟล์ **ข้อมูลสแกนนิ้ว** เช่น `26.11-25.12.2568.xlsx`  
       - สามารถอัปโหลดหลายไฟล์ได้ ถ้ามีหลายงวด  
    3. กดปุ่ม **Generate** เพื่อสร้างไฟล์ `template_filled.xlsx`  
    """
)

template_file = st.file_uploader("อัปโหลด template.xlsx", type=["xlsx"])
day_files_uploaded = st.file_uploader(
    "อัปโหลดไฟล์ข้อมูลสแกนนิ้ว (Day-by-Day) อย่างน้อย 1 ไฟล์ (.xlsx)",
    type=["xlsx"],
    accept_multiple_files=True,
)

if st.button("Generate"):
    if template_file is None:
        st.error("กรุณาอัปโหลด template.xlsx ก่อน")
    elif not day_files_uploaded:
        st.error("กรุณาอัปโหลดไฟล์ Day-by-Day อย่างน้อย 1 ไฟล์")
    else:
        with TemporaryDirectory() as tmpdir:
            # save template
            template_path = os.path.join(tmpdir, "template.xlsx")
            with open(template_path, "wb") as f:
                f.write(template_file.getbuffer())

            # save each day file to temp folder
            day_paths = []
            for uploaded in day_files_uploaded:
                fname = uploaded.name
                day_path = os.path.join(tmpdir, fname)
                with open(day_path, "wb") as f:
                    f.write(uploaded.getbuffer())
                day_paths.append(day_path)

            output_path = os.path.join(tmpdir, "template_filled.xlsx")

            try:
                with st.spinner("กำลังประมวลผล..."):
                    fill_template_from_days(template_path, day_paths, output_path)

                with open(output_path, "rb") as f:
                    data = f.read()

                st.success("สร้างไฟล์ template_filled.xlsx เรียบร้อยแล้ว 🎉")
                st.download_button(
                    label="⬇️ ดาวน์โหลด template_filled.xlsx",
                    data=data,
                    file_name="template_filled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"เกิดข้อผิดพลาดระหว่างการประมวลผล: {e}")
